import os
from typing import List, Dict, Any, Optional
from datetime import datetime
from io import BytesIO
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

from app.services.invoice_extractor import InvoiceInfo

logger = logging.getLogger(__name__)

class ExcelExporter:
    FIELD_LABELS = {
        "transaction_reference": ("交易编码", 28),
        "transaction_date": ("日期", 14),
        "receiver_account": ("收款账号/姓名", 20),
        "total_amount": ("金额", 15),
        "currency": ("货币", 8),
        "image_url": ("图片", 30),
        "validation_status": ("核对状态", 12),
        "needs_review": ("需人工核对", 12),
    }
    
    HEADER_FILL = PatternFill(
        start_color="4472C4",
        end_color="4472C4",
        fill_type="solid"
    )
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    IMAGE_SIZE = (200, 200)
    
    @staticmethod
    def create_workbook() -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "发票数据"
        return wb
    
    @staticmethod
    def setup_header(ws, fields: List[str]) -> None:
        for col, field in enumerate(fields, 1):
            header, width = ExcelExporter.FIELD_LABELS.get(field, (field, 12))
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = ExcelExporter.HEADER_FILL
            cell.font = ExcelExporter.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = ExcelExporter.BORDER
            ws.column_dimensions[get_column_letter(col)].width = width
        
        ws.row_dimensions[1].height = 25
    
    @staticmethod
    def add_invoice_row(
        ws,
        row_num: int,
        invoice_data: Dict[str, Any],
        fields: List[str],
        image_dir: str = None,
        include_images: bool = False
    ) -> None:
        for col, field in enumerate(fields, 1):
            if field == "image_url":
                value = ""
            else:
                value = invoice_data.get(field, "")
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = ExcelExporter.BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            if field == "total_amount":
                cell.number_format = '#,##0.00'
        
        ws.row_dimensions[row_num].height = 150
        
        image_url = invoice_data.get("image_url", "")
        if include_images and image_url and "image_url" in fields:
            if image_url.startswith("uploads") or image_url.startswith("/uploads"):
                image_path = image_url.replace("/", os.sep).replace("\\", os.sep)
                if image_path.startswith("uploads" + os.sep):
                    image_path = image_path[len("uploads" + os.sep):]
                if image_dir:
                    image_path = os.path.join(image_dir, image_path)
            else:
                image_path = os.path.join(image_dir, image_url) if image_dir else image_url
            
            if os.path.exists(image_path):
                try:
                    img = Image.open(image_path)
                    
                    temp_buffer = BytesIO()
                    img.save(temp_buffer, format='PNG')
                    temp_buffer.seek(0)
                    
                    xl_img = XLImage(temp_buffer)
                    xl_img.width = ExcelExporter.IMAGE_SIZE[0]
                    xl_img.height = ExcelExporter.IMAGE_SIZE[1]
                    
                    image_col_index = fields.index("image_url") + 1
                    cell_address = f"{get_column_letter(image_col_index)}{row_num}"
                    ws.add_image(xl_img, cell_address)
                except Exception as e:
                    logger.warning(f"Failed to embed image {image_path}: {e}")
    
    @staticmethod
    def export_to_file(
        invoices: List[Dict[str, Any]], 
        output_path: str,
        image_dir: str = None,
        include_summary: bool = True,
        include_images: bool = False
    ) -> str:
        wb = ExcelExporter.create_workbook()
        ws = wb.active
        
        fields = list(ExcelExporter.FIELD_LABELS.keys())
        ExcelExporter.setup_header(ws, fields)
        
        for idx, invoice in enumerate(invoices, 1):
            ExcelExporter.add_invoice_row(ws, idx + 1, invoice, fields, image_dir, include_images)
        
        if include_summary and invoices and "total_amount" in fields:
            summary_row = len(invoices) + 3
            ws.cell(row=summary_row, column=1, value="统计汇总")
            ws.cell(row=summary_row, column=1).font = Font(bold=True)
            
            total_amount = sum(
                float(inv.get("total_amount", 0) or 0) 
                for inv in invoices
            )
            
            ws.cell(row=summary_row + 1, column=1, value="记录数量:")
            ws.cell(row=summary_row + 1, column=2, value=len(invoices))
            ws.cell(row=summary_row + 2, column=1, value="金额合计:")
            ws.cell(row=summary_row + 2, column=2, value=total_amount)
            ws.cell(row=summary_row + 2, column=2).number_format = '#,##0.00'
        
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        wb.save(output_path)
        logger.info(f"Excel exported to {output_path}")
        
        return output_path
    
    @staticmethod
    def export_to_bytes(
        invoices: List[Dict[str, Any]],
        image_dir: str = None,
        include_images: bool = False
    ) -> BytesIO:
        wb = ExcelExporter.create_workbook()
        ws = wb.active
        
        fields = list(ExcelExporter.FIELD_LABELS.keys())
        if invoices:
            fields = [f for f in fields if f in invoices[0].keys()]
            if not fields:
                fields = list(ExcelExporter.FIELD_LABELS.keys())
        ExcelExporter.setup_header(ws, fields)
        
        for idx, invoice in enumerate(invoices, 1):
            ExcelExporter.add_invoice_row(ws, idx + 1, invoice, fields, image_dir, include_images)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def generate_filename(prefix: str = "发票数据") -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.xlsx"

excel_exporter = ExcelExporter()
